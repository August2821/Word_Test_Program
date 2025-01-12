import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import random

# 전역 변수
file_path = None  # 선택한 엑셀 파일의 경로
words = []  # 단어와 뜻의 리스트
word_length = 0  # 불러온 단어의 개수
current_question_index = 0  # 현재 문제 
correct_count = 0  # 맞춘 개수
wrong_count = 0  # 틀린 개수

def open_excel():
    """엑셀 파일을 선택하고 단어를 로드합니다."""
    global file_path  # 전역 변수 사용
    try:
        # 파일 선택 대화 상자 열기
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls"), ("All files", "*.*"))
        )
        if not file_path:  # 파일 선택을 취소한 경우
            return

        # 성공 메시지 표시
        messagebox.showinfo("성공", "엑셀 파일이 열렸습니다!")

        # 단어를 로드하는 함수 호출
        load_words()

    except Exception as e:
        # 파일 로드 실패 시 오류 메시지 출력
        messagebox.showerror("오류", f"엑셀 파일을 열 수 없습니다.\n{e}")

def load_words():
    """엑셀 파일에서 단어와 뜻을 로드합니다."""
    global file_path, words, word_length
    if not file_path:
        # 파일이 선택되지 않은 경우 경고 메시지 표시
        messagebox.showwarning("경고", "엑셀 파일을 먼저 선택하세요!")
        return

    try:
        # 엑셀 파일 열기
        workbook = load_workbook(file_path)
        sheet = workbook.active  # 첫 번째 시트 선택

        # 단어와 뜻을 튜플로 저장 (min_row=2는 헤더 제외)
        words = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True, min_row=2)]
        word_length = len(words)  # 단어 개수 저장

        # 단어가 4개 이하일 경우 경고 메시지 표시
        if word_length <= 4:
            messagebox.showwarning("경고", f"단어가 {word_length}개 입니다. 단어는 5개 이상이어야 합니다.")
            file_path = None  # 파일 경로 초기화

    except Exception as e:
        # 엑셀 데이터 로드 실패 시 오류 메시지 표시
        messagebox.showerror("오류", f"엑셀 데이터를 불러오는 중 오류 발생: {e}")

def start_test():
    """테스트 창을 열고 문제를 표시합니다."""
    global file_path, words, word_length, current_question_index, correct_count, wrong_count, show_correct_var, show_wrong_var
    if not file_path:
        # 파일이 선택되지 않은 경우 경고 메시지 표시
        messagebox.showwarning("경고", "엑셀 파일을 먼저 선택하세요!")
        return

    # 단어 리스트를 랜덤으로 섞음
    random.shuffle(words)

    correct_count = 0  # 맞춘 개수 초기화
    wrong_count = 0  # 틀린 개수 
    
    def update_counts():
        """맞은 개수와 틀린 개수를 업데이트합니다."""
        correct_label.config(text=f"맞은 개수: {correct_count}")
        wrong_label.config(text=f"틀린 개수: {wrong_count}")

    def next_question():
        """다음 문제를 화면에 표시합니다."""
        global current_question_index

        if current_question_index >= word_length:
            # 모든 문제가 끝났을 경우 완료 메시지 표시
            messagebox.showinfo("테스트 완료", "모든 문제가 출제되었습니다!")
            test_window.destroy()  # 테스트 창 닫기
            return

        # 현재 문제와 정답(뜻)을 설정
        current_word, correct_meaning = words[current_question_index]

        # 객관식 옵션 생성: 정답 + 랜덤 오답 3개
        meanings = list(set(word[1] for word in words))  # 중복 제거
        meanings.remove(correct_meaning)  # 정답 제거
        options = random.sample(meanings, 3) + [correct_meaning]
        random.shuffle(options)

        # 문제와 옵션 업데이트
        question_number_label.config(text=f"{current_question_index + 1}/{word_length}")
        question_label.config(text=f"{current_word}")
        for i, button in enumerate(option_buttons):
            button.config(text=f" {i + 1}. {options[i]}", command=lambda opt=options[i]: check_answer(opt, correct_meaning))

    def check_answer(selected_option, correct_meaning):
        """메시지 박스를 체크 했을 때, 사용자의 선택이 정답인지 확인합니다."""
        global current_question_index, correct_count, wrong_count, show_messagebox_var
        if show_messagebox_var.get():
            if selected_option == correct_meaning:
                correct_count += 1
                messagebox.showinfo("정답", "정답입니다!") # 정답 메시지 출력
            else:
                wrong_count += 1
                messagebox.showerror("오답", f"오답입니다! 정답: {correct_meaning}") # 오답 메시지 출력
        else:
            if selected_option == correct_meaning:
                correct_count += 1
            else:
                wrong_count += 1       

        # 다음 문제로 이동
        update_counts()
        current_question_index += 1
        next_question()

    # 테스트 창 생성
    test_window = tk.Toplevel(root)
    test_window.title("영어 단어 테스트")
    test_window.geometry("600x500")

    # 테스트 창의 그리드 설정 (중앙 배치를 위한 weight 조정)
    row_length = 7
    column_length = 3
    for i in range(row_length):
        test_window.grid_rowconfigure(i, weight=0)
    
    for i in range(column_length):
        test_window.grid_columnconfigure(i, weight=1)

    # 맞은 개수, 틀린 개수 라벨
    global correct_label, wrong_label
    correct_label = tk.Label(test_window, text=f"맞은 개수: {correct_count}", font=("Arial", 14))
    if show_correct_var.get():
        correct_label.grid(row=0, column=0, padx=10, pady=0, sticky="w")

    wrong_label = tk.Label(test_window, text=f"틀린 개수: {wrong_count}", font=("Arial", 14))
    if show_wrong_var.get():
        wrong_label.grid(row=1, column=0, padx=10, pady=0, sticky="w")

    # 문제 번호 라벨
    global question_number_label
    question_number_label = tk.Label(test_window, text="", font=("Arial", 16))
    question_number_label.grid(row=0, column=2, padx=10, pady=10, sticky="e")  # 그리드 중앙에 배치

    # 문제 텍스트 라벨
    global question_label
    question_label = tk.Label(test_window, text="", font=("Arial", 16))
    question_label.grid(row=1, column=1, pady=10)  # 그리드 중앙에 배치

    # 객관식 버튼 생성
    global option_buttons
    option_buttons = [tk.Button(test_window, font=("Arial", 14), width=30, anchor="w") for _ in range(4)]
    for i, button in enumerate(option_buttons):
        button.grid(row=i + 2, column=1, pady=5)  # 버튼을 중앙 열에 배치

    # 첫 번째 문제 표시
    current_question_index = 0
    next_question()

# Tkinter GUI 설정
root = tk.Tk()
root.title("영어 단어 테스트")
root.geometry("500x200")

# 버튼 추가
select_button = tk.Button(root, text="엑셀 파일 선택", command=open_excel, width=20, height=2)
select_button.grid(row=0, column=0, padx=10, pady=10, sticky="we")

'''
load_button = tk.Button(root, text="단어 불러오기", command=load_words, width=20, height=2)
load_button.grid(row=0, column=1, padx=10, pady=10, sticky="we")
'''

test_button = tk.Button(root, text="테스트 시작", command=start_test, width=20, height=2)
test_button.grid(row=0, column=1, padx=10, pady=10, sticky="we")

# 체크박스 상태를 저장하는 변수
show_correct_var = tk.BooleanVar(value=False)
show_wrong_var = tk.BooleanVar(value=False)
show_messagebox_var = tk.BooleanVar(value=False)
language_var = tk.StringVar(value="영어")

# 체크박스 추가
correct_checkbox = tk.Checkbutton(root, text="맞은 개수 보기", variable=show_correct_var)
correct_checkbox.grid(row=3, column=0, padx=10, pady=5, sticky="w")

wrong_checkbox = tk.Checkbutton(root, text="틀린 개수 보기", variable=show_wrong_var)
wrong_checkbox.grid(row=4, column=0, padx=10, pady=5, sticky="w")

messagebox_checkbox = tk.Checkbutton(root, text="정답 알림 보기", variable=show_messagebox_var)
messagebox_checkbox.grid(row=3, column=1, padx=10, pady=5, sticky="w")

# Tkinter 루프 시작
root.mainloop()
