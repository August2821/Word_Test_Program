import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import sqlite3
import os
import openpyxl
import random
from openpyxl import load_workbook, Workbook
from datetime import datetime
import smtplib
from email.message import EmailMessage

DB_NAME = 'student_db.db'
selected_file_path = None
file_checked_items = {}
file_id = {}

def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS word_test_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_path TEXT NOT NULL UNIQUE
        )
    ''')
    conn.commit()
    conn.close()

def load_word_test_files():
    file_checked_items.clear()
    file_id.clear()

    for row in file_tree.get_children():
        file_tree.delete(row)

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('SELECT id, file_path FROM word_test_files ORDER BY id')
    rows = cursor.fetchall()
    conn.close()

    for fid, file_path in rows:
        item_id = file_tree.insert('', 'end', values=('☐', file_path))
        file_checked_items[item_id] = False
        file_id[item_id] = fid

def toggle_checkbox(event):
    item_id = file_tree.identify_row(event.y)
    column = file_tree.identify_column(event.x)

    if column == '#1' and item_id:
        checked = file_checked_items.get(item_id, False)
        file_checked_items[item_id] = not checked
        new_icon = '☑' if not checked else '☐'
        file_tree.set(item_id, column='선택', value=new_icon)

def register_file():
    file_path = selected_file_path.get()
    if not file_path:
        messagebox.showwarning("경고", "먼저 파일을 선택하세요.")
        return

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM word_test_files WHERE file_path = ?", (file_path,))
    count = cursor.fetchone()[0]

    if count > 0:
        messagebox.showwarning("중복", "이미 등록된 파일입니다.")
    else:
        cursor.execute("INSERT INTO word_test_files (file_path) VALUES (?)", (file_path,))
        conn.commit()
        messagebox.showinfo("성공", "파일이 등록되었습니다.")
        selected_file_path.set("")
        uploaded_file_label.config(text="선택된 파일: 없음")
        load_word_test_files()

    conn.close()

def file_upload():
    file_path = filedialog.askopenfilename()
    if file_path:
        selected_file_path.set(file_path)
        uploaded_file_label.config(text=f"선택된 파일: {file_path}")

def cancel_upload():
    selected_file_path.set("")
    uploaded_file_label.config(text="선택된 파일: 없음")

# 선택된 파일 삭제
def delete_selected_files():
    to_delete_ids = [file_id[item] for item, checked in file_checked_items.items() if checked]
    if not to_delete_ids:
        messagebox.showwarning("경고", "삭제할 파일을 선택하세요.")
        return

    confirm = messagebox.askyesno("삭제 확인", f"{len(to_delete_ids)}개의 파일을 삭제하시겠습니까?")
    if not confirm:
        return

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.executemany("DELETE FROM word_test_files WHERE id = ?", [(fid,) for fid in to_delete_ids])
    conn.commit()
    conn.close()

    load_word_test_files()

# 파일 관리 셀 선택시 word_test() 실행
def handle_file_path_click(event):
    region = file_tree.identify("region", event.x, event.y)
    if region != "cell":
        return  # 헤더나 빈 공간 클릭 시 무시

    col = file_tree.identify_column(event.x)
    row = file_tree.identify_row(event.y)

    if col == "#2" and row:  # "#2"는 "파일 경로" 열
        file_path = file_tree.item(row, "values")[1]
        selected_file_path.set(file_path)
        word_test(file_path)

# =========================================== 파일 윈도우 ===========================================
def file_management():
    global selected_file_path, uploaded_file_label, file_tree

    init_db()

    file_window = tk.Tk()
    file_window.title("학생 프로그램")
    file_window.geometry("720x480")

    selected_file_path = tk.StringVar()

    # =========================================== 네비게이션 프레임 ===========================================
    nav_frame = tk.Frame(file_window)
    nav_frame.grid(row=0, column=0, padx=10, pady=(10, 0), sticky="ew")

    # tk.Button(nav_frame, text="홈으로", width=14).grid(row=0, column=0, padx=5)
    # tk.Button(nav_frame, text="종료", width=14, command=file_window.destroy).grid(row=0, column=1, padx=5)

    # tk.Button(nav_frame, text="홈으로", width=14).grid(row=0, column=0, padx=5)
    # tk.Button(nav_frame, text="종료", width=14, command=file_window.destroy).grid(row=0, column=1, padx=5)

    # 학생 이메일 표시 (왼쪽)
    student_email_label = tk.Label(nav_frame, text=f"학생 이메일: {student_email}", font=("Arial", 10))
    student_email_label.grid(row=1, column=0, padx=5, pady=(5, 10), sticky="w")

    # 선생님 이메일 표시 (오른쪽)
    teacher_email_label = tk.Label(nav_frame, text=f"선생님 이메일: {teacher_email}", font=("Arial", 10))
    teacher_email_label.grid(row=1, column=1, padx=5, pady=(5, 10), sticky="w")

    # =========================================== 상단 프레임 ===========================================
    top_frame = tk.LabelFrame(file_window, text="파일 선택")
    top_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

    uploaded_file_label = tk.Label(top_frame, text="선택된 파일: 없음", anchor="w")
    uploaded_file_label.grid(row=0, column=0, columnspan=2, padx=20, pady=(5, 0), sticky="w")

    tk.Button(top_frame, text="파일 업로드", command=file_upload, width=30, height=2)\
        .grid(row=1, column=0, columnspan=2, padx=20, pady=10, sticky="w")

    button_frame = tk.Frame(top_frame)
    button_frame.grid(row=2, column=0, columnspan=2, pady=10)

    tk.Button(button_frame, text="파일 등록", command=register_file, width=14, height=2)\
        .grid(row=0, column=0, padx=20)
    tk.Button(button_frame, text="취소", command=cancel_upload, width=14, height=2)\
        .grid(row=0, column=1, padx=20)

    # =========================================== 중간 프레임 ===========================================
    # 중간 프레임
    middle_frame = tk.LabelFrame(file_window, text="등록된 파일 목록")
    middle_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

    # 1. 위쪽 제어 프레임 (삭제 버튼 + 체크박스)
    top_controls_frame = tk.Frame(middle_frame)
    top_controls_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
    top_controls_frame.grid_columnconfigure(0, weight=1)
    top_controls_frame.grid_columnconfigure(1, weight=0)
    top_controls_frame.grid_columnconfigure(2, weight=0)

    # 삭제 버튼
    tk.Button(top_controls_frame, text="선택된 파일 삭제", command=delete_selected_files, width=20, height=2)\
        .grid(row=0, column=0, sticky="w")

    # 체크박스 상태 변수 선언 (전역)
    global show_correct_var, show_wrong_var
    show_correct_var = tk.BooleanVar(value=True)
    show_wrong_var = tk.BooleanVar(value=True)

    # 체크박스
    tk.Checkbutton(top_controls_frame, text="맞은 개수 보기", variable=show_correct_var)\
        .grid(row=0, column=1, padx=10, sticky="e")
    tk.Checkbutton(top_controls_frame, text="틀린 개수 보기", variable=show_wrong_var)\
        .grid(row=0, column=2, padx=10, sticky="e")

    # ▶️ 2. 트리뷰 프레임 (트리뷰 + 스크롤바)
    tree_frame = tk.Frame(middle_frame)
    tree_frame.grid(row=1, column=0, sticky="nsew")

    # 트리뷰
    columns = ("선택", "파일 경로")
    file_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
    file_tree.heading("선택", text="선택")
    file_tree.heading("파일 경로", text="파일 경로")
    file_tree.column("선택", width=60, anchor="center", stretch=False)
    file_tree.column("파일 경로", width=600, anchor="w", stretch=True)
    file_tree.grid(row=0, column=0, sticky="nsew")

    # 스크롤바
    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=file_tree.yview)
    file_tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.grid(row=0, column=1, sticky="ns")

    # 확장 설정
    file_window.grid_rowconfigure(2, weight=1)
    file_window.grid_columnconfigure(0, weight=1)
    middle_frame.grid_rowconfigure(1, weight=1)
    middle_frame.grid_columnconfigure(0, weight=1)
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)


    file_tree.bind("<Button-1>", toggle_checkbox)
    file_tree.bind("<ButtonRelease-1>", handle_file_path_click)

    load_word_test_files()
    file_window.mainloop()

# =========================================== 단어 테스트 윈도우 ===========================================
def word_test(file_path):
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        words = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # 단어와 뜻 모두 있을 때만 추가
                words.append((str(row[0]), str(row[1])))
        wb.close()
    except Exception as e:
        messagebox.showerror("오류", f"엑셀 파일을 열 수 없습니다:\n{e}")
        return

    if not words:
        messagebox.showwarning("경고", "엑셀 파일에 단어가 없습니다!")
        return

    # 테스트 변수들
    random.shuffle(words)
    total = len(words)
    current = [0]
    correct = [0]
    wrong = [0]
    wrong_words = []

    def check_answer(selected, correct_answer):
        if selected == correct_answer:
            correct[0] += 1
        else:
            wrong[0] += 1
            wrong_words.append((words[current[0]][0], correct_answer))

        # 점수 라벨 업데이트
        correct_label.config(text=f"맞은 개수: {correct[0]}")
        wrong_label.config(text=f"틀린 개수: {wrong[0]}")

        current[0] += 1
        if current[0] >= total:
            save_wrong_words(wrong_words)
            messagebox.showinfo("완료", f"테스트 종료!\n맞은 개수: {correct[0]}, 틀린 개수: {wrong[0]}")
            test_window.destroy()
        else:
            next_question()

    def next_question():
        word, correct_meaning = words[current[0]]
        question_label.config(text=f"{current[0] + 1}/{total}")
        word_label.config(text=word)

        all_meanings = list(set([w[1] for w in words if w[1] != correct_meaning]))
        options = random.sample(all_meanings, min(3, len(all_meanings))) + [correct_meaning]
        random.shuffle(options)
        for i in range(4):
            if i < len(options):
                option_buttons[i].config(text=options[i], state="normal",
                                         command=lambda opt=options[i]: check_answer(opt, correct_meaning))
            else:
                option_buttons[i].config(text="", state="disabled")

    def save_wrong_words(wrong_words):
        if not wrong_words:
            return

        try:
            # 현재 작업 디렉토리 기준 wrong_word 폴더 경로
            folder_path = os.path.join(os.getcwd(), "wrong_word")
            os.makedirs(folder_path, exist_ok=True)

            # 저장할 파일명: 틀린단어_YYYYMMDD_HHMMSS.xlsx
            now_str = datetime.now().strftime("%Y%m%d_%H%M%S")

            # 테스트 본 파일 이름
            original_name = os.path.splitext(os.path.basename(selected_file_path.get()))[0]
            filename = f"{original_name}_{now_str}.xlsx"
            save_path = os.path.join(folder_path, filename)

            wb = Workbook()
            sheet = wb.active
            sheet.title = "틀린 단어"
            sheet.append(["단어", "뜻"])
            for word, meaning in wrong_words:
                sheet.append([word, meaning])

            wb.save(save_path)
            messagebox.showinfo("저장 완료", f"틀린 단어가 자동 저장되었습니다:\n{save_path}")

            # 이메일 보내기 (선생님 이메일, 학생 이메일, 앱 비밀번호는 전역변수로 가정)
            try:
                msg = EmailMessage()
                msg["Subject"] = "틀린 단어 테스트 결과"
                msg["From"] = student_email
                msg["To"] = teacher_email
                msg.set_content(
                    f"맞은 개수: {correct[0]}\n"
                    f"틀린 개수: {wrong[0]}\n"
                    f"전체 개수: {total}\n"
                )

                # 첨부파일 추가
                with open(save_path, "rb") as f:
                    file_data = f.read()
                    file_name = os.path.basename(save_path)
                msg.add_attachment(file_data, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=file_name)

                # SMTP 서버 로그인 및 메일 전송
                with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                    smtp.login(student_email, app_password)
                    smtp.send_message(msg)

                messagebox.showinfo("메일 전송 완료", f"틀린 단어 파일이 {teacher_email}로 전송되었습니다.")
            except Exception as e:
                messagebox.showerror("메일 전송 실패", f"메일 전송 중 오류 발생:\n{e}")

        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    # UI 구성
    test_window = tk.Toplevel()
    test_window.title("단어 테스트")
    test_window.geometry("600x400")

    # 전체 레이아웃 구성
    test_window.grid_rowconfigure(0, weight=0)
    test_window.grid_rowconfigure(1, weight=0)
    test_window.grid_rowconfigure(2, weight=1)
    test_window.grid_columnconfigure(0, weight=1)

    # ------------------------ top_frame ------------------------
    top_frame = tk.Frame(test_window)
    top_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
    top_frame.grid_columnconfigure(0, weight=1)
    top_frame.grid_columnconfigure(1, weight=1)
    top_frame.grid_columnconfigure(2, weight=1)

    # 맞은 개수, 틀린 개수, 문제 번호
    correct_label = tk.Label(top_frame, text=f"맞은 개수: {correct[0]}", font=("Arial", 14))
    wrong_label = tk.Label(top_frame, text=f"틀린 개수: {wrong[0]}", font=("Arial", 14))
    question_label = tk.Label(top_frame, text="", font=("Arial", 14), anchor="e")

    if show_correct_var.get():
        correct_label.grid(row=0, column=0, sticky="w")
    if show_wrong_var.get():
        wrong_label.grid(row=0, column=1, sticky="w")

    question_label.grid(row=0, column=2, sticky="e")

    # ------------------------ middle_frame ------------------------
    middle_frame = tk.Frame(test_window)
    middle_frame.grid(row=1, column=0, pady=10)

    word_label = tk.Label(middle_frame, text="", font=("Arial", 24, "bold"))
    word_label.grid(row=0, column=0)

    # ------------------------ option_frame ------------------------
    option_frame = tk.Frame(test_window)
    option_frame.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="nsew")

    option_frame.grid_columnconfigure(0, weight=1)
    option_buttons = []
    for i in range(4):
        btn = tk.Button(option_frame, font=("Arial", 14), width=50)
        btn.grid(row=i, column=0, pady=5, sticky="ew")
        option_buttons.append(btn)

    next_question()

# =========================================== 로그인 윈도우 ===========================================
student_email = ""
app_password = ""
teacher_email = ""

def login():
    login_window = tk.Tk()
    login_window.title("Gmail 로그인")
    login_window.geometry("400x300")

    tk.Label(login_window, text="학생 Gmail 주소:", font=("Arial", 11)).grid(row=0, column=0, padx=10, pady=(20, 5), sticky="w")
    student_email_entry = tk.Entry(login_window, width=40)
    student_email_entry.grid(row=1, column=0, padx=10, pady=5, columnspan=2)

    tk.Label(login_window, text="앱 비밀번호:", font=("Arial", 11)).grid(row=2, column=0, padx=10, pady=5, sticky="w")
    password_entry = tk.Entry(login_window, width=40, show="*")
    password_entry.grid(row=3, column=0, padx=10, pady=5, columnspan=2)

    tk.Label(login_window, text="선생님 Gmail 주소:", font=("Arial", 11)).grid(row=4, column=0, padx=10, pady=5, sticky="w")
    teacher_email_entry = tk.Entry(login_window, width=40)
    teacher_email_entry.grid(row=5, column=0, padx=10, pady=5, columnspan=2)

    def do_login():
        global student_email, app_password, teacher_email

        email = student_email_entry.get().strip()
        password = password_entry.get().strip()
        teacher = teacher_email_entry.get().strip()

        if not email or not password or not teacher:
            messagebox.showwarning("입력 오류", "학생 이메일, 앱 비밀번호, 선생님 이메일을 모두 입력하세요.")
            return

        # 실제 로그인 시도
        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(email, password)
        except smtplib.SMTPAuthenticationError:
            messagebox.showerror("로그인 실패", "학생 이메일 또는 앱 비밀번호가 틀렸습니다.")
            return
        except Exception as e:
            messagebox.showerror("오류", f"알 수 없는 오류: {e}")
            return

        # 성공
        student_email = email
        app_password = password
        teacher_email = teacher
        login_window.destroy()
        file_management()

    tk.Button(login_window, text="로그인", command=do_login, width=12).grid(row=6, column=0, pady=20)

    login_window.grid_columnconfigure(0, weight=1)
    login_window.mainloop()

# =========================================== 실행 ===========================================
if __name__ == '__main__':
    file_management()